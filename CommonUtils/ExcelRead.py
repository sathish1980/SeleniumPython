import openpyxl as openpyxl


class Excel_Read:
    Dict ={}
    def read_data(self):
        book = openpyxl.load_workbook("C:\\Users\\sathishkumar\\PycharmProjects\\SeleniumFramework\\Input_File\\FB_Login.xlsx")
        sheet = book["input"]
        cell = sheet.cell(row=1, column=2)

        print(cell)
        sheet.cell(row=5, column=2).value= "newvalue"

        #for mar row
        max_row=sheet.max_row
        print(max_row)

        # for mar column
        max_column = sheet.max_column
        print(max_column)

        for i in range(1 , max_row+1):
            for j in range (1, max_column+1):
                print(sheet.cell(row=i, column=j).value)
    @staticmethod
    def retrundicvalur():
        Dict = {}
        book = openpyxl.load_workbook(
            "C:\\Users\\sathishkumar\\PycharmProjects\\SeleniumFramework\\Input_File\\FB_Login.xlsx")
        sheet = book["input"]
        cell = sheet.cell(row=1, column=2)

        print(cell)
        #sheet.cell(row=5, column=2).value = "newvalue"

        # for mar row
        max_row = sheet.max_row
        print(max_row)

        # for mar column
        max_column = sheet.max_column
        print(max_column)

        for i in range(2, max_row + 1):
            for j in range(2, max_column + 1):
                Dict[(sheet.cell(row=1, column=j).value)+str(i-1)]=sheet.cell(row=i, column=j).value

        return Dict

        #print(Dict)
#Excel_Read.retrundicvalur()
"""onj = Excel_Read()
onj.read_data()"""